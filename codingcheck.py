#KienTran
#Python37
#kientrancm@gmail.com


#lib
import os, re
import xlrd
import openpyxl as excel
from openpyxl import Workbook, load_workbook
import tkinter
from tkinter import *
from tkinter import filedialog

#Convert input file to xlsx for using
def cvt_xls_to_xlsx(src_file_path, name):
    book_xls = xlrd.open_workbook(src_file_path)
    book_xlsx = Workbook()
    base = os.path.basename(src_file_path)
    namefile = os.path.splitext(base)[0]
    sheet_names = book_xls.sheet_names()
    for sheet_index in range(0, len(sheet_names)):
        sheet_xls = book_xls.sheet_by_name(sheet_names[sheet_index])
        if sheet_index == 0:
            sheet_xlsx = book_xlsx.active
            sheet_xlsx.title = sheet_names[sheet_index]
        else:
            sheet_xlsx = book_xlsx.create_sheet(title=sheet_names[sheet_index])

        for row in range(0, sheet_xls.nrows):
            for col in range(0, sheet_xls.ncols):
                sheet_xlsx.cell(row=row + 1, column=col + 1).value = sheet_xls.cell_value(row, col)

    book_xlsx.save(name+".xlsx")

#GCAPE
def GCAPE_Check(nvm, coding, parameter):
    print("GCAPE Check")




## Find a in nvm file
def find_coding(name, default, nvm):
    nvmf = load_workbook(nvm)
    coding_hella = nvmf['CodingHella']
    coding_geely = nvmf['CodingGeely']
    coding_user = nvmf['CodingUser']

    coding_hella_flag = False
    coding_geely_flag = False
    coding_user_flag = False

    default_flag = False

    for datarows in range(2, coding_hella.max_row):
        if coding_hella.cell(row = datarows, column = 2).value == name:
            coding_hella_flag = True
            if coding_hella.cell(row = datarows, column = 4).value == default:
                default_flag = True
            else:
                default_flag = True
                new_default = coding_hella.cell(row = datarows, column = 4)
        else:
            pass
            #coding not in codinghellasheet

    if coding_hella_flag == False:
        for datarows in range(2, coding_geely.max_row):
            if coding_geely.cell(row = datarows, column = 2).value == name:
                coding_geely_flag = True
                if coding_geely.cell(row=datarows, column=4).value == default:
                    default_flag = True
                else:
                    default_flag = True
                    new_default = coding_geely.cell(row=datarows, column=4)
            else:
                pass
                # coding not in codinghellasheet
    elif coding_geely_flag == False:
        for datarows in range(2, coding_user.max_row):
            if coding_user.cell(row = datarows, column = 2).value == name:
                coding_user = True
                if coding_user.cell(row=datarows, column=4).value == default:
                    default_flag = True
                else:
                    default_flag = True
                    new_default = coding_user.cell(row=datarows, column=4)
            else:
                pass
                # coding not in codinghellasheet
    else:
        #not in coding




#----------------------------------------------------

#BCM
ws_coding = ("CodingHella", "CodingGeely", "CodingUser")

## Title of coding
Co_name_title = 2
Co_defaultvalue_title = 3
Co_SafetyClassification_F = 4
Co_ObjectContent_F = 5

## Title of parameter

def BCM_Check(nvm, coding, parameter):
    print("BCM Check")
    #open file nvm
    nvm_file = load_workbook(nvm)

    coding_file = load_workbook(coding)
    data_coding_file = coding_file.active
    max_row = data_coding_file.max_row
    max_column = data_coding_file.max_column

    data_coding_file.cell(row=1, column=11).value = "Default Value by Number"
    ##
    for data_rows in range(2,max_row + 1):
        req = data_coding_file.cell(row=data_rows, column = Co_ObjectContent_F).value
        if req == "Requirement":
            #Requirement then we will check the default value
            defaultvalue = data_coding_file.cell(row=data_rows, column = Co_defaultvalue_title).value

            #this column is the 1st value of default value in text
            defaultvalue_c1 = data_coding_file.cell(row=data_rows, column = 7).value

            # this column is the 2nd value of default value in text
            defaultvalue_c2 = data_coding_file.cell(row=data_rows, column = 9).value

            default_value = data_coding_file.cell(row=data_rows, column=11).value
            if defaultvalue == defaultvalue_c1:
                default_value = data_coding_file.cell(row=data_rows, column = 8).value
            elif defaultvalue == defaultvalue_c2:
                default_value = data_coding_file.cell(row=data_rows, column=10).value
            else:
                default_value = "NOK"
        else:
            data_coding_file.cell(row=data_rows, column=11).value = "NOT REQUIREMENT"

    coding_file.save(coding)
    coding_file.close()

    parameter_file = load_workbook(parameter)

#GUI
class GUI(tkinter.Frame):
    def __init__(self, root):
        tkinter.Frame.__init__(self, root)
        self.funtion = 0
        self.root = root
        self.initMenu()
        self.initGUI()

    def initMenu(self):
        self.root.title("CoPaCheck - Hella - created by KienTran")
        self.pack(fill=BOTH, expand=1)

        menubar = Menu(self.root)
        self.root.config(menu=menubar)

        #File menu
        fileMenu = Menu(menubar, tearoff=0)
        fileMenu.add_command(label="Exit", command=self.quit)
        menubar.add_cascade(label="File", menu=fileMenu)

        #Project Menu
        self.GCAPE = BooleanVar()
        self.BCM = BooleanVar()

        PrjMenu = Menu(menubar, tearoff=0)
        PrjMenu.add_checkbutton(label="GCAPE", onvalue=True, offvalue=False, variable = self.GCAPE, command=self._gcape)
        PrjMenu.add_checkbutton(label="BCM", onvalue=True, offvalue=False, variable = self.BCM, command=self._bcm)
        menubar.add_cascade(label="Project", menu=PrjMenu)

        #Help menu
        helpMenu = Menu(menubar, tearoff=0)
        helpMenu.add_command(label="Help")
        helpMenu.add_command(label="About")
        menubar.add_cascade(label="Help", menu=helpMenu)

    def _bcm(self):
        self.GCAPE.set(False)

    def _gcape(self):
        self.BCM.set(False)


    def initGUI(self):
        self.file_paths = []

        #Giao dien
        self.inputFrame = LabelFrame(self)
        self.optionFrame = LabelFrame(self)
        self.LeftBotLabelFrame = LabelFrame(self)
        self.RightBotLabelFrame = LabelFrame(self)

        self.inputFrame.grid(row=1, column=1, sticky = W)
        self.optionFrame.grid(row=1, column=2, sticky = W)
        self.LeftBotLabelFrame.grid(row=2, column=1, sticky = W)
        self.RightBotLabelFrame.grid(row=2, column=2, sticky = W)

        self.PathValue_NVM = tkinter.StringVar()
        self.PathValue_Coding = tkinter.StringVar()
        self.PathValue_Parameter = tkinter.StringVar()

        self.statusValue = StringVar()
        self.statusValue.set('Please select your option')
        self.NameVar = tkinter.StringVar()

        #input file
        self.InputGroupLabel = Label(self.inputFrame, text = "-----> Input files <-----", width = 60).grid(row = 1, column = 1, columnspan = 2)

        #Select NVM file
        self.ChooserButton = Button(self.inputFrame, width=10, bd=2, text='NVM file', command=self.OpenFile_NVM).grid(row=2, column=1)
        self.PathEntry = Entry(self.inputFrame, width=60, bd=2, textvariable=self.PathValue_NVM).grid(row=2, column=2)

        #Coding
        self.ChooserButton = Button(self.inputFrame, width=10, bd=2, text='Coding', command=self.OpenFile_Coding).grid(row=3, column=1)
        self.PathEntry = Entry(self.inputFrame, width=60, bd=2, textvariable=self.PathValue_Coding).grid(row=3, column=2)

        #Parameter
        self.ChooserButton = Button(self.inputFrame, width=10, bd=2, text='Parameter', command=self.OpenFile_Parameter).grid(row=4, column=1)
        self.PathEntry = Entry(self.inputFrame, width=60, bd=2, textvariable=self.PathValue_Parameter).grid(row=4, column=2)

        #run
        self.RunGroupLabel = Label(self.LeftBotLabelFrame, text='-----> Run <-----', width = 63).grid(row=1, column=1, columnspan = 2)
        self.generateButton = Button(self.LeftBotLabelFrame, text='Generate', command=self.MyGUI).grid(row=2, column=1, columnspan=2)
        self.statusLabel = Label(self.LeftBotLabelFrame, textvariable=self.statusValue).grid(row=3, column=1, columnspan=2)


    def OpenFile_NVM(self):
        file_paths = tkinter.filedialog.askopenfilename(filetype = (("NVM file", "*.xlsm"), ("All files", "*.*")), parent=self,)
        self.file_paths_nvm = file_paths
        self.PathValue_NVM.set(self.file_paths_nvm)

    def OpenFile_Coding(self):
        file_paths = tkinter.filedialog.askopenfilename(filetype = (("Coding file", "*.xlsx"), ("All files", "*.*")), parent=self,)
        self.file_paths_coding = file_paths
        self.PathValue_Coding.set(self.file_paths_coding)

    def OpenFile_Parameter(self):
        file_paths = tkinter.filedialog.askopenfilename(filetype = (("Parameter file", "*.xlsx"), ("All files", "*.*")), parent=self,)
        self.file_paths_parameter = file_paths
        self.PathValue_Parameter.set(self.file_paths_parameter)

    def setStatus(self, status):
        self.statusValue.set(status)

    def MyGUI(self):
        oktorun = False
        self.setStatus('Running...')
        if not str(self.PathValue_NVM.get()):
            self.setStatus('Missing NVM file')
        elif not str(self.PathValue_Coding.get()):
            self.setStatus('Missing Coding file')
        elif not str(self.PathValue_Parameter.get()):
            self.setStatus('Missing Parameter file')
        elif (self.GCAPE.get() == 0 and self.BCM.get() == 0):
            self.setStatus('Please select your project')
            oktorun = False
        else:
            oktorun = True
            HTMLlabel = 1

        if oktorun == True:
            self.setStatus("CoPaCheck is running....")
            #Xu ly duong dan file
            nvm_pfile = self.file_paths_nvm
            nvm_name = os.path.splitext(nvm_pfile)[0]
            cvt_xls_to_xlsx(nvm_pfile, nvm_name)
            nvm_pfile = nvm_name + ".xlsx"
            
            coding_pfile = self.file_paths_coding
            coding_name = os.path.splitext(coding_pfile)[0]

            parameter_pfile = self.file_paths_parameter
            parameter_name = os.path.splitext(parameter_pfile)[0]

            if self.GCAPE.get() == 1:
                #Perform GCAPE project
                print("gcape")
                #GCAPE_Check(nvm_pfile, coding_pfile, parameter_pfile)

            elif self.BCM.get() == 1:
                #Perform BCM project
                print("BCM")
                BCM_Check(nvm_pfile, coding_pfile, parameter_pfile)

            self.setStatus('Finished')


def runGUI():
    root = tkinter.Tk()
    rGUI = GUI(root)
    rGUI.pack()
    root.mainloop()

if __name__ == "__main__":
    runGUI()
