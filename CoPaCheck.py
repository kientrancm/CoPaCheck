#KienTran
#Python37
#kientrancm@gmail.com


#lib
import os, re
import xlrd
import openpyxl as excel
from openpyxl import Workbook, load_workbook
import tkinter
from tkinter import *
from tkinter import filedialog

#Convert input file to xlsx for using
def cvt_xls_to_xlsx(src_file_path, name):
    book_xls = xlrd.open_workbook(src_file_path)
    book_xlsx = Workbook()
    base = os.path.basename(src_file_path)
    namefile = os.path.splitext(base)[0]
    sheet_names = book_xls.sheet_names()
    for sheet_index in range(0, len(sheet_names)):
        sheet_xls = book_xls.sheet_by_name(sheet_names[sheet_index])
        if sheet_index == 0:
            sheet_xlsx = book_xlsx.active
            sheet_xlsx.title = sheet_names[sheet_index]
        else:
            sheet_xlsx = book_xlsx.create_sheet(title=sheet_names[sheet_index])

        for row in range(0, sheet_xls.nrows):
            for col in range(0, sheet_xls.ncols):
                sheet_xlsx.cell(row=row + 1, column=col + 1).value = sheet_xls.cell_value(row, col)

    book_xlsx.save(name+".xlsx")

def GCAPE_Coding(coding, nvm):
    print("Check coding..")
    nvmwb = load_workbook(nvm)
    coding_nvm = nvmwb['Coding']
    codingwb = load_workbook(coding)
    codingws = codingwb.active

    codingws.cell(row=1, column=5).value = "Status"
    codingws.cell(row=1, column=6).value = "Default value on NVM"

    for data_row in range(2, codingws.max_row + 1):
        coding_name = codingws.cell(row=data_row, column=2).value
        if coding_name[0:2] == "K_":
            checkisOK = False
            for i in range(2, coding_nvm.max_row+1):
                if coding_name == coding_nvm.cell(row=i, column=2).value:
                    checkisOK = True
                    if coding_nvm.cell(row=i, column=4).value == codingws.cell(row=data_row, column=3).value:
                        codingws.cell(row=data_row, column=5).value = "OK"
                        codingws.cell(row=data_row, column=6).value = coding_nvm.cell(row=i, column=4).value
                    else:
                        codingws.cell(row=data_row, column=5).value = "The default value is not OK"
                        codingws.cell(row=data_row, column=6).value = coding_nvm.cell(row=i, column=4).value

            if checkisOK == False:
                codingws.cell(row=data_row, column=5).value = "The coding not found in NVM"

    codingwb.save(coding)
    codingwb.close()
    nvmwb.close()

def GCAPE_Parameter(coding, nvm):
    print("Check parameter..")
    nvmwb = load_workbook(nvm)
    coding_nvm = nvmwb['Parameter']
    codingwb = load_workbook(coding)
    codingws = codingwb.active

    codingws.cell(row=1, column=5).value = "Status"
    codingws.cell(row=1, column=6).value = "Default value on NVM"
    codingws.cell(row=1, column=7).value = "Virtual Address"

    for data_row in range(2, codingws.max_row + 1):
        coding_name = codingws.cell(row=data_row, column=2).value
        if coding_name[0:4] == "PAR_":
            checkisOK = False
            for i in range(2, coding_nvm.max_row+1):
                if coding_name == coding_nvm.cell(row=i, column=2).value:
                    checkisOK = True
                    if int(float(coding_nvm.cell(row=i, column=4).value * coding_nvm.cell(row=i, column=9).value)) == codingws.cell(row=data_row, column=3).value:
                        codingws.cell(row=data_row, column=5).value = "OK"
                        codingws.cell(row=data_row, column=6).value = coding_nvm.cell(row=i, column=4).value
                        codingws.cell(row=data_row, column=7).value = coding_nvm.cell(row=i, column=8).value
                    else:
                        codingws.cell(row=data_row, column=5).value = "The default value is not OK"
                        codingws.cell(row=data_row, column=6).value = coding_nvm.cell(row=i, column=4).value
                        codingws.cell(row=data_row, column=7).value = coding_nvm.cell(row=i, column=8).value

            if checkisOK == False:
                codingws.cell(row=data_row, column=5).value = "The coding not found in NVM"

    codingwb.save(coding)
    codingwb.close()
    nvmwb.close()

#GCAPE
def GCAPE_Check(nvm, coding, parameter):
    print("GCAPE Check")
    GCAPE_Coding(coding, nvm)
    GCAPE_Parameter(parameter, nvm)
    print("Finished")



## Find a in nvm file
def find_coding(name, default, nvm):
    nvmf = load_workbook(nvm)
    coding_hella = nvmf['CodingHella']
    coding_geely = nvmf['CodingGeely']
    coding_user = nvmf['CodingUser']

    coding_hella_flag = False
    coding_geely_flag = False
    coding_user_flag = False

    default_flag = False

    for datarows in range(2, coding_hella.max_row):
        if coding_hella.cell(row = datarows, column = 2).value == name:
            coding_hella_flag = True
            if coding_hella.cell(row = datarows, column = 4).value == default:
                default_flag = True
            else:
                default_flag = True
                new_default = coding_hella.cell(row = datarows, column = 4)
        else:
            pass
            #coding not in codinghellasheet

    if coding_hella_flag == False:
        for datarows in range(2, coding_geely.max_row):
            if coding_geely.cell(row = datarows, column = 2).value == name:
                coding_geely_flag = True
                if coding_geely.cell(row=datarows, column=4).value == default:
                    default_flag = True
                else:
                    default_flag = True
                    new_default = coding_geely.cell(row=datarows, column=4)
            else:
                pass
                # coding not in codinghellasheet
    elif coding_geely_flag == False:
        for datarows in range(2, coding_user.max_row):
            if coding_user.cell(row = datarows, column = 2).value == name:
                coding_user = True
                if coding_user.cell(row=datarows, column=4).value == default:
                    default_flag = True
                else:
                    default_flag = True
                    new_default = coding_user.cell(row=datarows, column=4)
            else:
                pass
                # coding not in codinghellasheet
    else:
        pass
        #not in coding

#----------------------------------------------------
#DefaultValue from Text to number
#This functions is check a coding
#If the coding is requirement, we will convert the default values from text to number.
def Coding_DefaultValue(file):
    file_coding = load_workbook(file)
    ws_coding = file_coding.active
    max_row = ws_coding.max_row
    ws_coding.cell(row = 1, column=11).value = "Default value/Error"

    for data_rows in range(2, max_row + 1):
        req = ws_coding.cell(row=data_rows, column=5).value
        if req == "Requirement":
            defaultvalue = ws_coding.cell(row=data_rows, column=3).value
            if defaultvalue == ws_coding.cell(row=data_rows, column=7).value:
                ws_coding.cell(row=data_rows, column=11).value = ws_coding.cell(row=data_rows, column=8).value
            elif defaultvalue == ws_coding.cell(row=data_rows, column=9).value:
                ws_coding.cell(row=data_rows, column=11).value = ws_coding.cell(row=data_rows, column=10).value
            else:
                ws_coding.cell(row=data_rows, column=11).value = "NOK"
        else:
            ws_coding.cell(row=data_rows, column=11).value = ""

    file_coding.save(file)
    file_coding.close()

#This functions will be check the coding_name/coding file in coding/nvm file
#The output will be OK, NOK or Cannot find the coding in NVM
def BCM_Coding_Check(coding, nvm):
    print("Check coding...")
    nvm_wb = load_workbook(nvm)
    CodingHella = nvm_wb['CodingHella']
    CodingGeely = nvm_wb['CodingGeely']
    CodingUser = nvm_wb['CodingUser']

    Coding_DefaultValue(coding)

    coding_wb = load_workbook(coding)
    coding_ws = coding_wb.active
    max_row = coding_ws.max_row
    coding_ws.cell(row=1, column=12).value = "Status"
    coding_ws.cell(row=1, column=13).value = "Default value on NVM"

    #Get the coding name in coding file
    #and then check codingname in nvm file
    for data_rows in range(2, max_row + 1):
        req = coding_ws.cell(row=data_rows, column=5).value
        coding_name = coding_ws.cell(row=data_rows, column=2).value
        default_value = coding_ws.cell(row=data_rows, column=11).value
        if req == "Requirement":
            checkisOK = False
            for i in range(2, CodingHella.max_row + 1):
                if CodingHella.cell(row=i, column=2).value == coding_name:
                    checkisOK = True
                    if CodingHella.cell(row=i, column=4).value == default_value:
                        coding_ws.cell(row=data_rows, column=12).value = "OK"
                        coding_ws.cell(row=data_rows, column=13).value = default_value
                    else:
                        coding_ws.cell(row=data_rows, column=12).value = "The default value is not match. \nKindly help to check by manualy"
                        coding_ws.cell(row=data_rows, column=13).value = CodingHella.cell(row=i, column=4).value

            if checkisOK == False:
                for i in range(2, CodingGeely.max_row + 1):
                    if CodingGeely.cell(row=i, column=2).value == coding_name:
                        checkisOK = True
                        if CodingGeely.cell(row=i, column=4).value == default_value:
                            coding_ws.cell(row=data_rows, column=12).value = "OK"
                            coding_ws.cell(row=data_rows, column=13).value = default_value
                        else:
                            coding_ws.cell(row=data_rows, column=12).value = "The default value is not match. \nKindly help to check by manualy"
                            coding_ws.cell(row=data_rows, column=13).value = CodingGeely.cell(row=i, column=4).value

            if checkisOK == False:
                for i in range(2, CodingUser.max_row + 1):
                    if CodingUser.cell(row=i, column=2).value == coding_name:
                        checkisOK = True
                        if CodingUser.cell(row=i, column=4).value == default_value:
                            coding_ws.cell(row=data_rows, column=12).value = "OK"
                            coding_ws.cell(row=data_rows, column=13).value = default_value
                        else:
                            coding_ws.cell(row=data_rows, column=12).value = "The default value is not match. \nKindly help to check by manualy"
                            coding_ws.cell(row=data_rows, column=13).value = CodingUser.cell(row=i, column=4).value

            if checkisOK == False:
                coding_ws.cell(row=data_rows, column=12).value = "The coding name is not found in NVM. \nKindly help to check by manualy"

    coding_wb.save(coding)
    coding_wb.close()
    nvm_wb.close()

def BCM_Parameter_Check(parameter, nvm):
    print("Check parameter...")
    nvm_wb = load_workbook(nvm)
    ASILA = nvm_wb['ASILAParameter']
    ASILB = nvm_wb['ASILBParameter']
    NonASIL = nvm_wb['NonASILParameter']
    HWParameter = nvm_wb['HWParameter']

    parameter_wb = load_workbook(parameter)
    parameter_ws = parameter_wb.active
    max_row = parameter_ws.max_row

    parameter_ws.cell(row=1, column=7).value = "Status"
    parameter_ws.cell(row=1, column=8).value = "Address"
    parameter_ws.cell(row=1, column=9).value = "Default value on NVM"

    for data_rows in range(2, max_row + 1):
        #Check requirement
        req = parameter_ws.cell(row=data_rows, column=6).value
        par_name = parameter_ws.cell(row=data_rows, column=2).value
        par_default_value = parameter_ws.cell(row=data_rows, column=3).value
        par_resolution = parameter_ws.cell(row=data_rows, column=4).value
        par_safety = parameter_ws.cell(row=data_rows, column=5).value

        #value = par_default_value / par_resolution
        #if req == "Requirement":
        if par_name[0:4] == "p_n_" or par_name[0:4] == "p_t_":
            checkisOK = False
            #if par_safety == "ASIL A":
            for i in range(2, ASILA.max_row + 1):
                if ASILA.cell(row=i, column=2).value == par_name:
                    checkisOK = True
                    if ASILA.cell(row=i, column=4).value * par_resolution == par_default_value:
                        parameter_ws.cell(row=data_rows, column=7).value = "OK"
                        parameter_ws.cell(row=data_rows, column=8).value = ASILA.cell(row=i, column=8).value
                        parameter_ws.cell(row=data_rows, column=9).value = ASILA.cell(row=i, column=4).value
                    else:
                        parameter_ws.cell(row=data_rows, column=7).value = "The default value is NOK"
                        parameter_ws.cell(row=data_rows, column=8).value = ASILA.cell(row=i, column=8).value
                        parameter_ws.cell(row=data_rows, column=9).value = ASILA.cell(row=i, column=4).value
            #elif par_safety == "ASIL B":
            for i in range(2, ASILB.max_row + 1):
                if ASILB.cell(row=i, column=2).value == par_name:
                    checkisOK = True
                    if ASILB.cell(row=i, column=4).value * par_resolution == par_default_value:
                        parameter_ws.cell(row=data_rows, column=7).value = "OK"
                        parameter_ws.cell(row=data_rows, column=8).value = ASILB.cell(row=i, column=8).value
                        parameter_ws.cell(row=data_rows, column=9).value = ASILB.cell(row=i, column=4).value
                    else:
                        parameter_ws.cell(row=data_rows, column=7).value = "The default value is NOK"
                        parameter_ws.cell(row=data_rows, column=8).value = ASILB.cell(row=i, column=8).value
                        parameter_ws.cell(row=data_rows, column=9).value = ASILB.cell(row=i, column=4).value
            # elif par_safety == "QM":
            for i in range(2, NonASIL.max_row + 1):
                if NonASIL.cell(row=i, column=2).value == par_name:
                    checkisOK = True
                    if NonASIL.cell(row=i, column=4).value * par_resolution == par_default_value:
                        parameter_ws.cell(row=data_rows, column=7).value = "OK"
                        parameter_ws.cell(row=data_rows, column=8).value = NonASIL.cell(row=i, column=8).value
                        parameter_ws.cell(row=data_rows, column=9).value = NonASIL.cell(row=i, column=4).value
                    else:
                        parameter_ws.cell(row=data_rows, column=7).value = "The default value is NOK"
                        parameter_ws.cell(row=data_rows, column=8).value = NonASIL.cell(row=i, column=8).value
                        parameter_ws.cell(row=data_rows, column=9).value = NonASIL.cell(row=i, column=4).value
            # else:
            for i in range(2, HWParameter.max_row + 1):
                if HWParameter.cell(row=i, column=2).value == par_name:
                    checkisOK = True
                    if HWParameter.cell(row=i, column=4).value * par_resolution == par_default_value:
                        parameter_ws.cell(row=data_rows, column=7).value = "OK"
                        parameter_ws.cell(row=data_rows, column=8).value = HWParameter.cell(row=i, column=8).value
                        parameter_ws.cell(row=data_rows, column=9).value = HWParameter.cell(row=i, column=4).value
                    else:
                        parameter_ws.cell(row=data_rows, column=7).value = "The default value is NOK"
                        parameter_ws.cell(row=data_rows, column=8).value = HWParameter.cell(row=i, column=8).value
                        parameter_ws.cell(row=data_rows, column=9).value = HWParameter.cell(row=i, column=4).value

            if checkisOK == False:
                parameter_ws.cell(row=data_rows, column=7).value = "The parameter is not found in NVM"
                

    parameter_wb.save(parameter)
    parameter_wb.close()
    nvm_wb.close()

def BCM_Check(nvm, coding, parameter):
    BCM_Coding_Check(coding, nvm)
    BCM_Parameter_Check(parameter, nvm)
    print("Finished!")

#GUI
class GUI(tkinter.Frame):
    def __init__(self, root):
        tkinter.Frame.__init__(self, root)
        self.funtion = 0
        self.root = root
        self.initMenu()
        self.initGUI()

    def initMenu(self):
        self.root.title("CoPaCheck - Hella - created by KienTran")
        self.pack(fill=BOTH, expand=1)

        menubar = Menu(self.root)
        self.root.config(menu=menubar)

        #File menu
        fileMenu = Menu(menubar, tearoff=0)
        fileMenu.add_command(label="Exit", command=self.quit)
        menubar.add_cascade(label="File", menu=fileMenu)

        #Project Menu
        self.GCAPE = BooleanVar()
        self.BCM = BooleanVar()

        PrjMenu = Menu(menubar, tearoff=0)
        PrjMenu.add_checkbutton(label="GCAPE", onvalue=True, offvalue=False, variable = self.GCAPE, command=self._gcape)
        PrjMenu.add_checkbutton(label="BCM", onvalue=True, offvalue=False, variable = self.BCM, command=self._bcm)
        menubar.add_cascade(label="Project", menu=PrjMenu)

        #Help menu
        helpMenu = Menu(menubar, tearoff=0)
        helpMenu.add_command(label="Help")
        helpMenu.add_command(label="About")
        menubar.add_cascade(label="Help", menu=helpMenu)

    def _bcm(self):
        self.GCAPE.set(False)

    def _gcape(self):
        self.BCM.set(False)


    def initGUI(self):
        self.file_paths = []

        #Giao dien
        self.inputFrame = LabelFrame(self)
        self.optionFrame = LabelFrame(self)
        self.LeftBotLabelFrame = LabelFrame(self)
        self.RightBotLabelFrame = LabelFrame(self)

        self.inputFrame.grid(row=1, column=1, sticky = W)
        self.optionFrame.grid(row=1, column=2, sticky = W)
        self.LeftBotLabelFrame.grid(row=2, column=1, sticky = W)
        self.RightBotLabelFrame.grid(row=2, column=2, sticky = W)

        self.PathValue_NVM = tkinter.StringVar()
        self.PathValue_Coding = tkinter.StringVar()
        self.PathValue_Parameter = tkinter.StringVar()

        self.statusValue = StringVar()
        self.statusValue.set('Please select your option')
        self.NameVar = tkinter.StringVar()

        #input file
        self.InputGroupLabel = Label(self.inputFrame, text = "-----> Input files <-----", width = 60).grid(row = 1, column = 1, columnspan = 2)

        #Select NVM file
        self.ChooserButton = Button(self.inputFrame, width=10, bd=2, text='NVM file', command=self.OpenFile_NVM).grid(row=2, column=1)
        self.PathEntry = Entry(self.inputFrame, width=60, bd=2, textvariable=self.PathValue_NVM).grid(row=2, column=2)

        #Coding
        self.ChooserButton = Button(self.inputFrame, width=10, bd=2, text='Coding', command=self.OpenFile_Coding).grid(row=3, column=1)
        self.PathEntry = Entry(self.inputFrame, width=60, bd=2, textvariable=self.PathValue_Coding).grid(row=3, column=2)

        #Parameter
        self.ChooserButton = Button(self.inputFrame, width=10, bd=2, text='Parameter', command=self.OpenFile_Parameter).grid(row=4, column=1)
        self.PathEntry = Entry(self.inputFrame, width=60, bd=2, textvariable=self.PathValue_Parameter).grid(row=4, column=2)

        #run
        self.RunGroupLabel = Label(self.LeftBotLabelFrame, text='-----> Run <-----', width = 63).grid(row=1, column=1, columnspan = 2)
        self.generateButton = Button(self.LeftBotLabelFrame, text='Generate', command=self.MyGUI).grid(row=2, column=1, columnspan=2)
        self.statusLabel = Label(self.LeftBotLabelFrame, textvariable=self.statusValue).grid(row=3, column=1, columnspan=2)


    def OpenFile_NVM(self):
        file_paths = tkinter.filedialog.askopenfilename(filetype = (("NVM file", "*.xlsm"), ("All files", "*.*")), parent=self,)
        self.file_paths_nvm = file_paths
        self.PathValue_NVM.set(self.file_paths_nvm)

    def OpenFile_Coding(self):
        file_paths = tkinter.filedialog.askopenfilename(filetype = (("Coding file", "*.xlsx"), ("All files", "*.*")), parent=self,)
        self.file_paths_coding = file_paths
        self.PathValue_Coding.set(self.file_paths_coding)

    def OpenFile_Parameter(self):
        file_paths = tkinter.filedialog.askopenfilename(filetype = (("Parameter file", "*.xlsx"), ("All files", "*.*")), parent=self,)
        self.file_paths_parameter = file_paths
        self.PathValue_Parameter.set(self.file_paths_parameter)

    def setStatus(self, status):
        self.statusValue.set(status)

    def MyGUI(self):
        oktorun = False
        self.setStatus('Running...')
        if not str(self.PathValue_NVM.get()):
            self.setStatus('Missing NVM file')
        elif not str(self.PathValue_Coding.get()):
            self.setStatus('Missing Coding file')
        elif not str(self.PathValue_Parameter.get()):
            self.setStatus('Missing Parameter file')
        elif (self.GCAPE.get() == 0 and self.BCM.get() == 0):
            self.setStatus('Please select your project')
            oktorun = False
        else:
            oktorun = True
            HTMLlabel = 1

        if oktorun == True:
            self.setStatus("CoPaCheck is running....")
            #Xu ly duong dan file
            nvm_pfile = self.file_paths_nvm
            nvm_name = os.path.splitext(nvm_pfile)[0]
            cvt_xls_to_xlsx(nvm_pfile, nvm_name)
            nvm_pfile = nvm_name + ".xlsx"
            
            coding_pfile = self.file_paths_coding
            coding_name = os.path.splitext(coding_pfile)[0]

            parameter_pfile = self.file_paths_parameter
            parameter_name = os.path.splitext(parameter_pfile)[0]

            if self.GCAPE.get() == 1:
                #Perform GCAPE project
                GCAPE_Check(nvm_pfile, coding_pfile, parameter_pfile)

            elif self.BCM.get() == 1:
                #Perform BCM project
                print("BCM")
                BCM_Check(nvm_pfile, coding_pfile, parameter_pfile)

            self.setStatus('Finished')


def runGUI():
    root = tkinter.Tk()
    rGUI = GUI(root)
    rGUI.pack()
    root.mainloop()

if __name__ == "__main__":
    runGUI()
